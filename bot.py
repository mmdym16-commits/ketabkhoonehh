import telebot
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from openpyxl import load_workbook
import random
from collections import Counter

# ========== توکن ربات (اینجا بذار) ==========
TOKEN = "8919993932:AAHZZQ1oLnKlQBQ_f5rlNkUkKKxhFXY5I5c"
bot = telebot.TeleBot(TOKEN)

user_data = {}

# ========== بارگذاری دیتابیس با openpyxl ==========
def load_database():
    try:
        wb = load_workbook("database.xlsx")
        ws = wb.active
        data = []
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                data.append(dict(zip(headers, row)))
        print(f"✅ {len(data)} مطلب بارگذاری شد!")
        return data
    except Exception as e:
        print(f"❌ خطا: {e}")
        return []

db = load_database()

# ========== منوی اصلی ==========
def get_main_menu():
    keyboard = InlineKeyboardMarkup(row_width=2)
    buttons = [
        InlineKeyboardButton("🔍 جستجو", callback_data="search"),
        InlineKeyboardButton("🏷️ تگ‌ها", callback_data="tag_search"),
        InlineKeyboardButton("📂 دسته‌بندی", callback_data="categories"),
        InlineKeyboardButton("🎲 تصادفی", callback_data="random")
    ]
    keyboard.add(*buttons)
    return keyboard

# ========== ارسال فایل ==========
def send_media(chat_id, row):
    file_id = row.get('file_id')
    file_type = row.get('file_type')
    
    caption = f"📖 *{row.get('title', 'بدون عنوان')}*\n\n"
    if row.get('content'):
        caption += f"{row.get('content')}\n\n"
    caption += f"🏷️ تگ‌ها: {row.get('tags', '')}\n"
    caption += f"📂 دسته: {row.get('category', '')}"
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("❤️ ذخیره", callback_data=f"save_{row.get('id', 0)}"))
    
    if file_type == "photo":
        bot.send_photo(chat_id, file_id, caption=caption, parse_mode='Markdown', reply_markup=keyboard)
    elif file_type == "video":
        bot.send_video(chat_id, file_id, caption=caption, parse_mode='Markdown', reply_markup=keyboard)
    elif file_type == "document":
        bot.send_document(chat_id, file_id, caption=caption, parse_mode='Markdown', reply_markup=keyboard)
    elif file_type == "audio":
        bot.send_audio(chat_id, file_id, caption=caption, parse_mode='Markdown', reply_markup=keyboard)
    elif file_type == "animation":
        bot.send_animation(chat_id, file_id, caption=caption, parse_mode='Markdown', reply_markup=keyboard)
    else:
        bot.send_message(chat_id, caption, parse_mode='Markdown', reply_markup=keyboard)

# ========== جستجو ==========
def search_db(keyword):
    results = []
    for item in db:
        if (keyword.lower() in item.get('title', '').lower() or
            keyword.lower() in item.get('content', '').lower() or
            keyword.lower() in item.get('tags', '').lower()):
            results.append(item)
    return results

# ========== صفحه‌بندی ==========
def show_results_page(chat_id, results, page=0):
    if not results:
        bot.send_message(chat_id, "❌ نتیجه‌ای پیدا نشد!")
        return
    
    total_pages = (len(results) - 1) // 3 + 1
    if page >= total_pages:
        page = total_pages - 1
    if page < 0:
        page = 0
    
    start = page * 3
    end = min(start + 3, len(results))
    current_page = results[start:end]
    
    user_data[chat_id] = {'results': results, 'page': page}
    
    for row in current_page:
        send_media(chat_id, row)
    
    if total_pages > 1:
        nav_keyboard = InlineKeyboardMarkup()
        nav_buttons = []
        if page > 0:
            nav_buttons.append(InlineKeyboardButton("⬅️ قبلی", callback_data=f"page_{page-1}"))
        nav_buttons.append(InlineKeyboardButton(f"📄 {page+1}/{total_pages}", callback_data="none"))
        if page < total_pages - 1:
            nav_buttons.append(InlineKeyboardButton("➡️ بعدی", callback_data=f"page_{page+1}"))
        nav_keyboard.add(*nav_buttons)
        bot.send_message(chat_id, "📌 صفحات:", reply_markup=nav_keyboard) 
        # ========== دریافت file_id ==========
@bot.message_handler(content_types=['photo', 'video', 'document', 'audio', 'animation'])
def get_file_id(message):
    if message.photo:
        file_id = message.photo[-1].file_id
        f_type = "photo"
    elif message.video:
        file_id = message.video.file_id
        f_type = "video"
    elif message.document:
        file_id = message.document.file_id
        f_type = "document"
    elif message.audio:
        file_id = message.audio.file_id
        f_type = "audio"
    elif message.animation:
        file_id = message.animation.file_id
        f_type = "animation"
    
    bot.reply_to(
        message,
        f"✅ file_id:\n\n{file_id}\n\nنوع: {f_type}",
        parse_mode='Markdown'
    )

# ========== اجرا ==========
if name == "__main__":
    print("🤖 ربات روشن شد!")
    bot.infinity_polling()
    # ========== دستور start ==========
@bot.message_handler(commands=['start'])
def start(message):
    bot.send_message(
        message.chat.id,
        f"🧪 *به ربات علمی خوش اومدی!*\n📚 {len(db)} مطلب داریم.",
        parse_mode='Markdown',
        reply_markup=get_main_menu()
    )

# ========== جستجوی آزاد ==========
@bot.message_handler(func=lambda msg: True)
def handle_search(message):
    keyword = message.text.strip()
    if len(keyword) < 2:
        bot.reply_to(message, "❌ حداقل ۲ کاراکتر وارد کن!")
        return
    
    results = search_db(keyword)
    
    if not results:
        bot.reply_to(message, "❌ مطلبی پیدا نشد!")
        return
    
    bot.reply_to(message, f"✅ {len(results)} نتیجه:")
    show_results_page(message.chat.id, results)

# ========== دکمه‌ها ==========
@bot.callback_query_handler(func=lambda call: True)
def handle_callback(call):
    chat_id = call.message.chat.id
    data = call.data
    bot.answer_callback_query(call.id)
    
    if data == "search":
        bot.edit_message_text("🔍 کلمه مورد نظر رو بفرست:", chat_id, call.message.message_id)
    
    elif data == "tag_search":
        all_tags = []
        for item in db:
            if item.get('tags'):
                all_tags.extend([t.strip() for t in str(item['tags']).split(',')])
        top_tags = Counter(all_tags).most_common(15)
        
        keyboard = InlineKeyboardMarkup(row_width=2)
        for tag, count in top_tags:
            keyboard.add(InlineKeyboardButton(f"#{tag} ({count})", callback_data=f"tag_{tag}"))
        keyboard.add(InlineKeyboardButton("🔙 برگشت", callback_data="back"))
        bot.edit_message_text("🏷️ تگ‌ها:", chat_id, call.message.message_id, reply_markup=keyboard)
    
    elif data == "categories":
        categories = list(set([item.get('category') for item in db if item.get('category')]))
        keyboard = InlineKeyboardMarkup(row_width=2)
        for cat in categories:
            count = len([item for item in db if item.get('category') == cat])
            keyboard.add(InlineKeyboardButton(f"📁 {cat} ({count})", callback_data=f"cat_{cat}"))
        keyboard.add(InlineKeyboardButton("🔙 برگشت", callback_data="back"))
        bot.edit_message_text("📂 دسته‌بندی:", chat_id, call.message.message_id, reply_markup=keyboard)
    
    elif data == "random":
        if not db:
            bot.edit_message_text("❌ مطلبی نیست!", chat_id, call.message.message_id)
            return
        row = random.choice(db)
        bot.delete_message(chat_id, call.message.message_id)
        send_media(chat_id, row)
    
    elif data == "back":
        bot.edit_message_text("🧪 منوی اصلی:", chat_id, call.message.message_id, reply_markup=get_main_menu())
    
    elif data.startswith("tag_"):
        tag = data.replace("tag_", "")
        results = [item for item in db if tag.lower() in item.get('tags', '').lower()]
        if not results:
            bot.edit_message_text(f"❌ #{tag} پیدا نشد!", chat_id, call.message.message_id)
            return
        bot.edit_message_text(f"🏷️ #{tag}: {len(results)} مطلب", chat_id, call.message.message_id)
        show_results_page(chat_id, results)
    
    elif data.startswith("cat_"):
        category = data.replace("cat_", "")
        results = [item for item in db if item.get('category') == category]
        bot.edit_message_text(f"📂 {category}: {len(results)} مطلب", chat_id, call.message.message_id)
        show_results_page(chat_id, results)
    
    elif data.startswith("page_"):
        page = int(data.replace("page_", ""))
        if chat_id in user_data:
            results = user_data[chat_id]['results']
            show_results_page(chat_id, results, page)
    
    elif data.startswith("save_"):
        bot.answer_callback_query(call.id, "❤️ ذخیره شد!", show_alert=False)